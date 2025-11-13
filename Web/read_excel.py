#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import sys
import os


try:
    import pandas as pd
    print("Using pandas to read Excel file...")

    # 读取所有工作表
    if len(sys.argv) < 2:
        print("Usage: python read_excel.py <path_to_excel_file>")
        sys.exit(1)
    excel_file = sys.argv[1]
    df_dict = pd.read_excel(excel_file, sheet_name=None)

    for sheet_name, df in df_dict.items():
        print(f"\n=== Sheet: {sheet_name} ===")
        print(f"Shape: {df.shape}")
        print("\nColumns:")
        for i, col in enumerate(df.columns):
            print(f"  {i+1}. {col}")

        print("\nData:")
        print(df.to_string(index=False))
        print("=" * 80)

except ImportError:
    print("pandas not available, trying openpyxl...")
    try:
        import openpyxl

        if len(sys.argv) < 2:
            print("Usage: python read_excel.py <path_to_excel_file>")
            sys.exit(1)
        excel_file = sys.argv[1]
        wb = openpyxl.load_workbook(excel_file)

        for ws in wb.worksheets:
            print(f"\n=== Sheet: {ws.title} ===")
            print(f"Max row: {ws.max_row}, Max column: {ws.max_column}")

            # 读取所有数据
            data = []
            for row in ws.iter_rows(values_only=True):
                data.append(row)

            if data:
                # 打印表头
                print("\nHeaders:")
                if data[0]:
                    for i, header in enumerate(data[0]):
                        if header is not None:
                            print(f"  {i+1}. {header}")

                # 打印所有数据
                print("\nData:")
                for i, row in enumerate(data):
                    if any(cell is not None for cell in row):
                        print(f"Row {i+1}: {[str(cell) if cell is not None else '' for cell in row]}")

            print("=" * 80)

    except ImportError:
        print("Neither pandas nor openpyxl available. Cannot read Excel file.")
        sys.exit(1)

except Exception as e:
    print(f"Error reading Excel file: {e}")
    sys.exit(1)
